#!/usr/bin/env python3
"""
JPX Market Analysis - Data Extractor (extract.py)
==================================================
Extracts data from JPX Excel files and outputs structured JSON.

Usage:
    python extract.py [--dir DATA_DIR] [--out data.json] [--nikkei CLOSE] [--vi VI]

Inputs (auto-detected by filename pattern):
    Daily (required):
        YYYYMMDDopen_interest.xlsx
        YYYYMMDD_volume_by_participant_whole_day_J-NET.xlsx
        YYYYMMDD_market_data_whole_day.xlsx
    Weekly (optional, Monday update):
        YYYYMMDD_nk225op_oi_by_tp.xlsx
        YYYYMMDD_indexfut_oi_by_tp.xlsx
    Weekly (optional, Thursday update):
        stock_vol_1_YYMMDD.xls

Output:
    data.json with sections s01..s11 + metadata
"""

import argparse
import glob
import json
import math
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl

# ============================================================
# Constants
# ============================================================

MAJOR_MONTHS = [3, 6, 9, 12]

PARTICIPANT_RULES = {
    'us': ['ゴールドマン', 'ＪＰモルガン', 'シティ', 'ビーオブエー', 'モルガン'],
    'eu': ['ＵＢＳ', 'ソシエテ', 'バークレイズ', 'ＨＳＢＣ', 'ドイツ', 'ナティクシス'],
    'hf': ['ＡＢＮ', 'ＢＮＰ', 'サスケハナ', 'インタラクティブ', 'フィリップ'],
    'domestic': [
        '野村', '大和', 'みずほ', '三菱', 'ＳＢＩ', '楽天', '松井', '岩井',
        'ＳＭＢＣ', '豊証券', '立花', 'むさし', '日産証券', '岡三', '安藤',
        '光世', '東海東京', '広田', '三田', '極東', '三晃'
    ],
}

PARTICIPANT_LABELS = {
    'us': '米系', 'eu': '欧系', 'hf': 'HF代理', 'domestic': '国内', 'other': 'その他'
}

WEEKDAY_JA = ['月', '火', '水', '木', '金', '土', '日']


# ============================================================
# Helpers
# ============================================================

def safe_num(v, default=0):
    """Convert cell value to number safely."""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace(',', '').replace(' ', '').strip()
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return default


def classify_participant(name):
    """Classify participant into us/eu/hf/domestic/other."""
    if not name:
        return 'other'
    s = str(name)
    for cat, keywords in PARTICIPANT_RULES.items():
        for kw in keywords:
            if kw in s:
                return cat
    return 'other'


def is_overseas(cat):
    return cat in ('us', 'eu', 'hf')


def next_major_month(dt):
    """Find the nearest upcoming major expiry month (3/6/9/12)."""
    y, m = dt.year, dt.month
    for mm in MAJOR_MONTHS:
        if mm >= m:
            return y, mm
    return y + 1, MAJOR_MONTHS[0]


def sq_date(y, m):
    """Calculate SQ date = 2nd Friday of month."""
    first = datetime(y, m, 1)
    dow = first.weekday()  # 0=Mon
    first_fri = first + timedelta(days=(4 - dow) % 7)
    return first_fri + timedelta(days=7)


def business_days_between(d1, d2):
    """Count business days between d1 and d2 (exclusive of d1, inclusive of d2)."""
    count = 0
    cur = d1 + timedelta(days=1)
    while cur <= d2:
        if cur.weekday() < 5:
            count += 1
        cur += timedelta(days=1)
    return count


def norm_cdf(x):
    """Cumulative normal distribution (Abramowitz-Stegun approximation)."""
    if x >= 0:
        t = 1.0 / (1.0 + 0.2316419 * x)
    else:
        t = 1.0 / (1.0 - 0.2316419 * x)
    d = 0.3989422804014327  # 1/sqrt(2*pi)
    poly = ((((1.330274429 * t - 1.821255978) * t + 1.781477937) * t
             - 0.356563782) * t + 0.319381530) * t
    if x >= 0:
        return 1.0 - d * math.exp(-0.5 * x * x) * poly
    else:
        return d * math.exp(-0.5 * x * x) * poly


def otm_probability(forward, strike, sigma, T, option_type='put'):
    """Calculate probability of OTM expiry under BS."""
    if T <= 0 or sigma <= 0:
        return 0.5
    d2 = math.log(forward / strike) / (sigma * math.sqrt(T))
    if option_type == 'put':
        return norm_cdf(d2)
    else:  # call
        return norm_cdf(-d2)


def bs_price(opt_type, K, F, sigma, T):
    """Black-Scholes option price (European, no discounting for simplicity)."""
    if T <= 0 or sigma <= 0:
        if opt_type == 'put':
            return max(K - F, 0)
        return max(F - K, 0)
    sqrtT = math.sqrt(T)
    d1 = (math.log(F / K) + 0.5 * sigma * sigma * T) / (sigma * sqrtT)
    d2 = d1 - sigma * sqrtT
    if opt_type == 'put':
        return K * norm_cdf(-d2) - F * norm_cdf(-d1)
    else:
        return F * norm_cdf(d1) - K * norm_cdf(d2)


def round500(x):
    """Round to nearest 500."""
    return round(x / 500) * 500


# ============================================================
# File detection
# ============================================================

def detect_files(data_dir):
    """Auto-detect JPX files in directory by naming pattern.
    When multiple files match the same type, picks the one with the latest date.
    """
    files = {}
    all_files = os.listdir(data_dir)

    # Collect candidates per type: {type: [(date_str, filepath), ...]}
    candidates = {
        'open_interest': [],
        'jnet': [],
        'market_data': [],
        'op_participants': [],
        'fut_participants': [],
        'stock_vol': [],
    }

    for f in all_files:
        fp = os.path.join(data_dir, f)
        fl = f.lower()

        # Extract date from filename
        m = re.search(r'(\d{8})', f)
        if not m:
            m = re.search(r'(\d{6})', f)  # YYMMDD for stock_vol
        fdate = m.group(1) if m else '00000000'

        if 'open_interest' in fl and fl.endswith('.xlsx'):
            candidates['open_interest'].append((fdate, fp))

        elif 'volume_by_participant' in fl and 'j-net' in fl and fl.endswith('.xlsx'):
            candidates['jnet'].append((fdate, fp))

        elif 'market_data' in fl and fl.endswith('.xlsx'):
            candidates['market_data'].append((fdate, fp))

        elif 'nk225op_oi_by_tp' in fl and fl.endswith('.xlsx'):
            candidates['op_participants'].append((fdate, fp))

        elif 'indexfut_oi_by_tp' in fl and fl.endswith('.xlsx'):
            candidates['fut_participants'].append((fdate, fp))

        elif (fl.startswith('stock_vol') or fl.startswith('stock_val')) and fl.endswith('.xls'):
            candidates['stock_vol'].append((fdate, fp))

    # Pick latest file per type
    for key, cands in candidates.items():
        if cands:
            cands.sort(key=lambda x: x[0], reverse=True)  # latest date first
            files[key] = cands[0][1]

    # Extract analysis date from the latest daily file (open_interest preferred)
    for key in ['open_interest', 'jnet', 'market_data']:
        cands = candidates.get(key, [])
        if cands:
            cands.sort(key=lambda x: x[0], reverse=True)
            date_candidate = cands[0][0]
            if len(date_candidate) == 8:
                files['date_str'] = date_candidate
                break

    return files


# ============================================================
# Section Extractors
# ============================================================

def extract_atm(wb_market):
    """Extract ATM (futures settlement price) from market_data."""
    ws = wb_market['指数先物']
    atm = None
    today = datetime.now()
    major_y, major_m = next_major_month(today)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value) if row[0].value else ''
        if '日経225' in a_val and 'mini' not in a_val.lower():
            # Scan rows below for the major month settlement
            for r2 in ws.iter_rows(min_row=row[0].row + 1,
                                   max_row=min(row[0].row + 20, ws.max_row),
                                   values_only=False):
                b_val = str(r2[1].value) if len(r2) > 1 and r2[1].value else ''
                # Look for major month like "2026年06月限" or just month numbers
                if any(str(major_m).zfill(2) in b_val for _ in [1]):
                    # Q column (index 16) = settlement price
                    q_val = safe_num(r2[16].value if len(r2) > 16 else None)
                    if q_val and q_val > 10000:
                        atm = int(q_val)
                        break
            if atm:
                break

    # Fallback: scan for any settlement price in reasonable range
    if not atm:
        for row in ws.iter_rows(min_row=30, max_row=70, values_only=False):
            a_val = str(row[0].value) if row[0].value else ''
            if '日経225' in a_val and 'mini' not in a_val.lower():
                for r2 in ws.iter_rows(min_row=row[0].row + 1,
                                       max_row=min(row[0].row + 15, ws.max_row),
                                       values_only=False):
                    q_val = safe_num(r2[16].value if len(r2) > 16 else None)
                    if q_val and 20000 < q_val < 80000:
                        atm = int(q_val)
                        break
            if atm:
                break

    return atm


def extract_ohlc_pivot(wb_market):
    """Extract N225 futures OHLC and compute Pivot Points."""
    ws = wb_market['指数先物']
    result = {}

    # Find 取引相場表 section, then N225 large row
    in_table = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''

        if '取引相場表' in a_val:
            in_table = True
            continue

        if not in_table:
            continue

        # Find N225 large (not mini, not micro)
        if '日経225' in a_val and 'mini' not in a_val.lower() and 'ミニ' not in a_val and 'マイクロ' not in a_val:
            # Next rows have the data. Scan for the major month.
            for r2 in ws.iter_rows(min_row=row[0].row, max_row=min(row[0].row + 5, ws.max_row), values_only=False):
                b_val = str(r2[1].value).strip() if len(r2) > 1 and r2[1].value else ''
                # Look for 6-digit month code like 202606
                if not b_val or not b_val.isdigit():
                    continue

                # Full-day OHLC from 3 sessions
                # D(3)=夜間始値 E(4)=夜間高値 F(5)=夜間安値 G(6)=夜間終値
                # H(7)=前場始値 I(8)=前場高値 J(9)=前場安値 K(10)=前場終値
                # L(11)=後場始値 M(12)=後場高値 N(13)=後場安値 O(14)=後場終値
                # Q(16)=清算値
                night_o = safe_num(r2[3].value if len(r2) > 3 else None)
                night_h = safe_num(r2[4].value if len(r2) > 4 else None)
                night_l = safe_num(r2[5].value if len(r2) > 5 else None)
                am_h = safe_num(r2[8].value if len(r2) > 8 else None)
                am_l = safe_num(r2[9].value if len(r2) > 9 else None)
                pm_h = safe_num(r2[12].value if len(r2) > 12 else None)
                pm_l = safe_num(r2[13].value if len(r2) > 13 else None)
                pm_c = safe_num(r2[14].value if len(r2) > 14 else None)
                settle = safe_num(r2[16].value if len(r2) > 16 else None)

                # Compute full-day OHLC
                highs = [h for h in [night_h, am_h, pm_h] if h and h > 10000]
                lows = [l for l in [night_l, am_l, pm_l] if l and l > 10000]

                if highs and lows and night_o and night_o > 10000:
                    O = int(night_o)
                    H = int(max(highs))
                    L = int(min(lows))
                    C = int(settle) if settle and settle > 10000 else int(pm_c) if pm_c and pm_c > 10000 else 0

                    if C > 0:
                        result['open'] = O
                        result['high'] = H
                        result['low'] = L
                        result['close'] = C
                        result['range'] = H - L
                        result['month'] = b_val

                        # Pivot Points
                        PP = round((H + L + C) / 3)
                        R1 = 2 * PP - L
                        S1 = 2 * PP - H
                        R2 = PP + (H - L)
                        S2 = PP - (H - L)
                        R3 = H + 2 * (PP - L)
                        S3 = L - 2 * (H - PP)

                        result['pivot'] = PP
                        result['r1'] = R1
                        result['r2'] = R2
                        result['r3'] = R3
                        result['s1'] = S1
                        result['s2'] = S2
                        result['s3'] = S3

                        return result

    return result


def extract_vi_from_excel(wb_market):
    """Extract Nikkei VI futures settlement price from market_data."""
    ws = wb_market['指数先物']
    in_table = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''

        if '取引相場表' in a_val:
            in_table = True
            continue

        if not in_table:
            continue

        # Find 日経VI row
        if '日経VI' in a_val or '日経ＶＩ' in a_val:
            # Q column (index 16) = settlement price
            q_val = safe_num(row[16].value if len(row) > 16 else None)
            if q_val and 10 < q_val < 100:
                return q_val

    return None


def extract_s01(nikkei_close, vi, atm):
    """Section ①: Nikkei/VI analysis + predicted ranges."""
    data = {
        'nikkei_close': nikkei_close,
        'vi': vi,
    }

    if nikkei_close and vi:
        vi_dec = vi / 100.0
        range_1d = nikkei_close * vi_dec / math.sqrt(250)
        range_1w = nikkei_close * vi_dec / math.sqrt(52)
        data['range_1d'] = {
            'width': round(range_1d),
            'low': round(nikkei_close - range_1d),
            'high': round(nikkei_close + range_1d),
        }
        data['range_1w'] = {
            'width': round(range_1w),
            'low': round(nikkei_close - range_1w),
            'high': round(nikkei_close + range_1w),
        }

    return data


def extract_s02(wb_oi):
    """Section ②: Futures open interest changes.
    
    Sheet layout: left side (A-F) and right side (H-M) have different products.
    指数先物 section starts with '＜OSE 指数先物取引＞'.
    Left: 日経225 large (rows ~30-49), TOPIX (rows ~50-63)
    Right: 日経225mini (rows ~36-52), ミニTOPIX (rows ~58-61)
    """
    ws = wb_oi['デリバティブ建玉残高状況']
    result = {'nk225_large': {}, 'nk225_mini': {}, 'topix': {}}

    # Step 1: Find 指数先物 section boundaries
    fut_start = None
    fut_end = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        if '指数先物取引' in a_val:
            fut_start = row[0].row
        elif fut_start and ('商品先物' in a_val or '国債先物オプション' in a_val or '指数オプション' in a_val):
            fut_end = row[0].row
            break
    
    if not fut_start:
        return result
    if not fut_end:
        fut_end = min(fut_start + 80, ws.max_row)

    # Step 2: Parse LEFT side (A-F) — 日経225 large, then TOPIX
    left_section = None
    for row in ws.iter_rows(min_row=fut_start, max_row=fut_end, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        b_val = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''

        # Detect left-side product headers
        if '日経225' in a_val and 'mini' not in a_val.lower() and 'ミニ' not in a_val and 'マイクロ' not in a_val and 'オプション' not in a_val:
            left_section = 'nk225_large'
        elif 'TOPIX' in a_val and 'ミニ' not in a_val:
            left_section = 'topix'
        elif a_val and a_val not in ('', ' ') and left_section and '日経' not in a_val and 'TOPIX' not in a_val:
            # Hit a different product on left side — stop
            if left_section == 'topix':
                left_section = None

        if not left_section:
            continue

        # Total row
        if '合計' in b_val:
            result[left_section]['total_oi'] = safe_num(row[3].value if len(row) > 3 else None)
            result[left_section]['total_change'] = safe_num(row[4].value if len(row) > 4 else None)
            if left_section == 'nk225_large':
                left_section = None  # done with large, TOPIX will follow

        # Month rows
        elif b_val and ('月限' in b_val or '年' in b_val):
            oi = safe_num(row[3].value if len(row) > 3 else None)
            chg = safe_num(row[4].value if len(row) > 4 else None)
            if 'months' not in result[left_section]:
                result[left_section]['months'] = []
            result[left_section]['months'].append({
                'label': b_val, 'oi': oi, 'change': chg,
            })

    # Step 3: Parse RIGHT side (H-M) — 日経225mini
    right_section = None
    for row in ws.iter_rows(min_row=fut_start, max_row=fut_end, values_only=False):
        h_val = str(row[7].value).strip() if len(row) > 7 and row[7].value else ''
        i_val = str(row[8].value).strip() if len(row) > 8 and row[8].value else ''

        # Detect right-side product headers
        if '日経225mini' in h_val or '日経225ミニ' in h_val:
            right_section = 'nk225_mini'
        elif '日経225マイクロ' in h_val or 'ミニTOPIX' in h_val:
            right_section = None  # skip micro/mini-TOPIX for now

        if right_section != 'nk225_mini':
            continue

        # Total row
        if '合計' in i_val:
            result['nk225_mini']['total_oi'] = safe_num(row[10].value if len(row) > 10 else None)
            result['nk225_mini']['total_change'] = safe_num(row[11].value if len(row) > 11 else None)
            right_section = None  # done

        # Month rows
        elif i_val and ('月限' in i_val or '年' in i_val):
            oi = safe_num(row[10].value if len(row) > 10 else None)
            chg = safe_num(row[11].value if len(row) > 11 else None)
            if 'months' not in result['nk225_mini']:
                result['nk225_mini']['months'] = []
            result['nk225_mini']['months'].append({
                'label': i_val, 'oi': oi, 'change': chg,
            })

    return result


def extract_s03(wb_market):
    """Section ③: Options trading volume/value.
    
    Sheet layout: two blocks side by side.
    Block 1 (rows ~6-16): Left=国債先物OP, Right=日経225OP (N-X columns)
    Block 2 (rows ~18-28): Left=日経225ミニOP (B-L cols), Right=JPX400OP (N-X cols)
    Block 3 (rows ~30-40): Left=TOPIXオプション, Right=有価証券OP
    Each block has 合計 row (total) followed by J-NET row (next row below).
    """
    ws = wb_market['オプション']
    result = {'large': {}, 'mini': {}}

    large_done = False
    mini_done = False
    in_mini_block = False

    for row in ws.iter_rows(min_row=1, max_row=60, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        b_val = str(row[1].value).strip() if len(row) > 1 and row[1].value else ''

        # Detect mini options block start
        if '日経225ミニ' in b_val or 'ミニオプション' in b_val:
            in_mini_block = True

        if '合' not in a_val or '計' not in a_val:
            continue

        # First 合計 before mini block = large options (right side: N-X)
        if not in_mini_block and not large_done:
            result['large']['put_volume'] = safe_num(row[13].value if len(row) > 13 else None)
            result['large']['put_value'] = safe_num(row[15].value if len(row) > 15 else None)
            result['large']['call_volume'] = safe_num(row[17].value if len(row) > 17 else None)
            result['large']['call_value'] = safe_num(row[19].value if len(row) > 19 else None)
            result['large']['total_volume'] = safe_num(row[21].value if len(row) > 21 else None)
            result['large']['total_value'] = safe_num(row[23].value if len(row) > 23 else None)
            # J-NET = next row
            nr = row[0].row + 1
            for r2 in ws.iter_rows(min_row=nr, max_row=nr, values_only=False):
                result['large']['jnet_volume'] = safe_num(r2[21].value if len(r2) > 21 else None)
                result['large']['jnet_value'] = safe_num(r2[23].value if len(r2) > 23 else None)
            large_done = True

        # First 合計 after mini block header = mini options (left side: B-L)
        elif in_mini_block and not mini_done:
            result['mini']['put_volume'] = safe_num(row[1].value if len(row) > 1 else None)
            result['mini']['put_value'] = safe_num(row[3].value if len(row) > 3 else None)
            result['mini']['call_volume'] = safe_num(row[5].value if len(row) > 5 else None)
            result['mini']['call_value'] = safe_num(row[7].value if len(row) > 7 else None)
            result['mini']['total_volume'] = safe_num(row[9].value if len(row) > 9 else None)
            result['mini']['total_value'] = safe_num(row[11].value if len(row) > 11 else None)
            # J-NET = next row
            nr = row[0].row + 1
            for r2 in ws.iter_rows(min_row=nr, max_row=nr, values_only=False):
                result['mini']['jnet_volume'] = safe_num(r2[9].value if len(r2) > 9 else None)
                result['mini']['jnet_value'] = safe_num(r2[11].value if len(r2) > 11 else None)
            mini_done = True

        if large_done and mini_done:
            break

    # Calculate J-NET ratios
    for block in ('large', 'mini'):
        d = result[block]
        tv = d.get('total_value', 0)
        jv = d.get('jnet_value', 0)
        d['jnet_ratio'] = round(jv / tv * 100, 1) if tv else 0

    return result


def extract_s04(wb_oi):
    """Section ④: Options OI changes (large + mini)."""
    result = {'large': {}, 'mini': {}}

    # --- Large (別紙1) ---
    ws1 = wb_oi['別紙1']
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''

        if 'プット合計' in a_val:
            result['large']['put_total_change'] = safe_num(row[3].value if len(row) > 3 else None)
            result['large']['put_total_oi'] = safe_num(row[2].value if len(row) > 2 else None)

        if 'コール合計' in g_val:
            result['large']['call_total_change'] = safe_num(row[9].value if len(row) > 9 else None)
            result['large']['call_total_oi'] = safe_num(row[8].value if len(row) > 8 else None)

        # Major month subtotals
        if '限月合計' in a_val:
            if 'put_month_totals' not in result['large']:
                result['large']['put_month_totals'] = []
            result['large']['put_month_totals'].append({
                'label': a_val,
                'oi': safe_num(row[2].value if len(row) > 2 else None),
                'change': safe_num(row[3].value if len(row) > 3 else None),
            })

        if '限月合計' in g_val:
            if 'call_month_totals' not in result['large']:
                result['large']['call_month_totals'] = []
            result['large']['call_month_totals'].append({
                'label': g_val,
                'oi': safe_num(row[8].value if len(row) > 8 else None),
                'change': safe_num(row[9].value if len(row) > 9 else None),
            })

    # --- Mini (別紙2) ---
    if '別紙2' in wb_oi.sheetnames:
        ws2 = wb_oi['別紙2']
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
            a_val = str(row[0].value).strip() if row[0].value else ''
            g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''

            if 'プット合計' in a_val:
                result['mini']['put_total_change'] = safe_num(row[3].value if len(row) > 3 else None)
                result['mini']['put_total_oi'] = safe_num(row[2].value if len(row) > 2 else None)

            if 'コール合計' in g_val:
                result['mini']['call_total_change'] = safe_num(row[9].value if len(row) > 9 else None)
                result['mini']['call_total_oi'] = safe_num(row[8].value if len(row) > 8 else None)

            # Weekly month subtotals
            if '限月合計' in a_val:
                if 'put_month_totals' not in result['mini']:
                    result['mini']['put_month_totals'] = []
                result['mini']['put_month_totals'].append({
                    'label': a_val,
                    'oi': safe_num(row[2].value if len(row) > 2 else None),
                    'change': safe_num(row[3].value if len(row) > 3 else None),
                })

            if '限月合計' in g_val:
                if 'call_month_totals' not in result['mini']:
                    result['mini']['call_month_totals'] = []
                result['mini']['call_month_totals'].append({
                    'label': g_val,
                    'oi': safe_num(row[8].value if len(row) > 8 else None),
                    'change': safe_num(row[9].value if len(row) > 9 else None),
                })

    return result


def extract_s05(wb_oi):
    """Section ⑤: Important OI changes (|change| >= 300)."""
    ws = wb_oi['別紙1']
    changes = []

    # Regex for contract names
    pat_put = re.compile(r'NIKKEI\s*225\s*P\s*(\d{4})-(\d+)')
    pat_call = re.compile(r'NIKKEI\s*225\s*C\s*(\d{4})-(\d+)')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        # Check puts (A column)
        a_val = str(row[0].value).strip() if row[0].value else ''
        m = pat_put.match(a_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            chg = safe_num(row[3].value if len(row) > 3 else None)
            oi = safe_num(row[2].value if len(row) > 2 else None)
            if abs(chg) >= 300:
                changes.append({
                    'type': 'P',
                    'expiry': expiry,
                    'strike': strike,
                    'change': chg,
                    'oi': oi,
                    'name': a_val,
                })

        # Check calls (G column)
        g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''
        m = pat_call.match(g_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            chg = safe_num(row[9].value if len(row) > 9 else None)
            oi = safe_num(row[8].value if len(row) > 8 else None)
            if abs(chg) >= 300:
                changes.append({
                    'type': 'C',
                    'expiry': expiry,
                    'strike': strike,
                    'change': chg,
                    'oi': oi,
                    'name': g_val,
                })

    # Sort: expiry -> P before C -> strike ascending
    changes.sort(key=lambda x: (x['expiry'], 0 if x['type'] == 'P' else 1, x['strike']))
    return changes


def extract_s06(wb_oi, atm):
    """Section ⑥: OI distribution by expiry month (500-yen intervals, ATM ± 5000)."""
    if not atm:
        return {'error': 'ATM not determined'}

    ws = wb_oi['別紙1']

    atm_round = round500(atm)
    low = atm_round - 5000
    high = atm_round + 5000

    # Parse: expiry -> strike -> {put_oi, put_chg, call_oi, call_chg}
    pat_put = re.compile(r'NIKKEI\s*225\s*P\s*(\d{4})-(\d+)')
    pat_call = re.compile(r'NIKKEI\s*225\s*C\s*(\d{4})-(\d+)')

    # {expiry: {strike: {put_oi, put_change, call_oi, call_change}}}
    by_expiry = defaultdict(lambda: defaultdict(lambda: {'put_oi': 0, 'put_change': 0, 'call_oi': 0, 'call_change': 0}))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''

        # Put side
        m = pat_put.match(a_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            if low <= strike <= high and strike % 500 == 0:
                oi = safe_num(row[2].value if len(row) > 2 else None)
                chg = safe_num(row[3].value if len(row) > 3 else None)
                by_expiry[expiry][strike]['put_oi'] += oi
                by_expiry[expiry][strike]['put_change'] += chg

        # Call side
        m = pat_call.match(g_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            if low <= strike <= high and strike % 500 == 0:
                oi = safe_num(row[8].value if len(row) > 8 else None)
                chg = safe_num(row[9].value if len(row) > 9 else None)
                by_expiry[expiry][strike]['call_oi'] += oi
                by_expiry[expiry][strike]['call_change'] += chg

    # Sort expiries: nearest first
    sorted_expiries = sorted(by_expiry.keys())

    # Build per-expiry distributions
    expiry_distributions = []
    for expiry in sorted_expiries:
        data = by_expiry[expiry]
        # Skip expiries with very little OI
        total_oi = sum(d['put_oi'] + d['call_oi'] for d in data.values())
        if total_oi < 100:
            continue

        # Expiry label: '2604' -> '2026年04月限'
        label = '20%s年%s月限' % (expiry[:2], expiry[2:]) if len(expiry) == 4 else expiry

        dist = []
        for strike in range(low, high + 1, 500):
            d = data.get(strike, {'put_oi': 0, 'put_change': 0, 'call_oi': 0, 'call_change': 0})
            dist.append({
                'strike': strike,
                'put_oi': d['put_oi'],
                'put_change': d['put_change'],
                'call_oi': d['call_oi'],
                'call_change': d['call_change'],
                'is_atm': (strike == atm_round),
            })
        expiry_distributions.append({
            'expiry': expiry,
            'label': label,
            'total_oi': total_oi,
            'distribution': dist,
        })

    # Also build combined (all expiries) for backward compatibility
    combined = []
    for strike in range(low, high + 1, 500):
        p_oi = sum(by_expiry[e].get(strike, {}).get('put_oi', 0) for e in by_expiry)
        p_chg = sum(by_expiry[e].get(strike, {}).get('put_change', 0) for e in by_expiry)
        c_oi = sum(by_expiry[e].get(strike, {}).get('call_oi', 0) for e in by_expiry)
        c_chg = sum(by_expiry[e].get(strike, {}).get('call_change', 0) for e in by_expiry)
        combined.append({
            'strike': strike,
            'put_oi': p_oi,
            'put_change': p_chg,
            'call_oi': c_oi,
            'call_change': c_chg,
            'is_atm': (strike == atm_round),
        })

    return {
        'atm': atm,
        'atm_round': atm_round,
        'distribution': combined,  # backward compat for Max Pain etc
        'by_expiry': expiry_distributions,
    }


def extract_s07(wb_jnet):
    """Section ⑦: J-NET large trades (NIKKEI 225 OOP)."""
    ws = wb_jnet['手口上位一覧']
    trades = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        c_val = str(row[2].value).strip() if len(row) > 2 and row[2].value else ''

        if 'NIKKEI 225 OOP' not in c_val:
            continue

        f_val = str(row[5].value).strip() if len(row) > 5 and row[5].value else ''
        h_val = safe_num(row[7].value if len(row) > 7 else None)

        # Exclude SBI
        if 'ＳＢＩ' in f_val or 'SBI' in f_val.upper():
            continue

        # Filter >= 100
        if h_val < 100:
            continue

        trades.append({
            'product': c_val,
            'participant': f_val,
            'volume': h_val,
            'category': classify_participant(f_val),
        })

    # Detect matching pairs (same product, same volume, 2 participants)
    from collections import Counter
    pair_key = lambda t: (t['product'], t['volume'])
    counts = Counter(pair_key(t) for t in trades)
    for t in trades:
        if counts[pair_key(t)] == 2:
            t['is_pair'] = True
        else:
            t['is_pair'] = False

    # Sort by product, then volume descending
    trades.sort(key=lambda t: (t['product'], -t['volume']))

    return trades


def extract_s09_futures(wb_fut):
    """Section ⑨-A: Futures participant positions."""
    # First sheet (typically only one)
    ws = wb_fut[wb_fut.sheetnames[0]]
    sections = {}
    current_section = None
    current_data = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''

        # Detect section headers
        if '＜日経225先物＞' in a_val or '＜日経225＞' in a_val:
            if 'mini' not in a_val.lower() and 'ミニ' not in a_val:
                current_section = 'nk225_large'
                current_data = {'sellers': [], 'buyers': []}
                sections[current_section] = current_data
                continue

        if '＜日経225mini＞' in a_val.lower() or '＜日経225ミニ＞' in a_val:
            current_section = 'nk225_mini'
            current_data = {'sellers': [], 'buyers': []}
            sections[current_section] = current_data
            continue

        if '＜TOPIX先物＞' in a_val or '＜TOPIX＞' in a_val:
            current_section = 'topix'
            current_data = {'sellers': [], 'buyers': []}
            sections[current_section] = current_data
            continue

        if not current_data:
            continue

        # D=seller name (index 3), E=seller volume (index 4)
        # G=buyer name (index 6), H=buyer volume (index 7)
        d_val = str(row[3].value).strip() if len(row) > 3 and row[3].value else ''
        e_val = safe_num(row[4].value if len(row) > 4 else None)
        g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''
        h_val = safe_num(row[7].value if len(row) > 7 else None)

        if d_val and e_val > 0:
            current_data['sellers'].append({
                'name': d_val,
                'volume': e_val,
                'category': classify_participant(d_val),
            })

        if g_val and h_val > 0:
            current_data['buyers'].append({
                'name': g_val,
                'volume': h_val,
                'category': classify_participant(g_val),
            })

    # Compute overseas vs domestic net for each section
    for sec_name, sec_data in sections.items():
        overseas_sell = sum(s['volume'] for s in sec_data['sellers'] if is_overseas(s['category']))
        domestic_sell = sum(s['volume'] for s in sec_data['sellers'] if not is_overseas(s['category']))
        overseas_buy = sum(b['volume'] for b in sec_data['buyers'] if is_overseas(b['category']))
        domestic_buy = sum(b['volume'] for b in sec_data['buyers'] if not is_overseas(b['category']))
        sec_data['overseas_net'] = overseas_buy - overseas_sell
        sec_data['domestic_net'] = domestic_buy - domestic_sell

    return sections


def extract_s09_options(wb_op):
    """Section ⑨-B: Options participant positions."""
    ws = wb_op[wb_op.sheetnames[0]]

    # Aggregate: participant -> {put_sell, put_buy, call_sell, call_buy}
    participants = defaultdict(lambda: {
        'put_sell': 0, 'put_buy': 0, 'call_sell': 0, 'call_buy': 0,
        'put_sell_detail': [], 'put_buy_detail': [],
        'call_sell_detail': [], 'call_buy_detail': [],
    })

    current_strike = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        # B column (index 1) = put strike anchor
        b_val = row[1].value if len(row) > 1 else None
        if b_val is not None:
            try:
                strike_candidate = int(float(str(b_val).replace(',', '')))
                if 20000 < strike_candidate < 80000:
                    current_strike = strike_candidate
            except (ValueError, TypeError):
                pass

        if not current_strike:
            continue

        # Put sell: D=name(3), E=volume(4)
        d_val = str(row[3].value).strip() if len(row) > 3 and row[3].value else ''
        e_val = safe_num(row[4].value if len(row) > 4 else None)
        if d_val and e_val > 0:
            participants[d_val]['put_sell'] += e_val
            participants[d_val]['put_sell_detail'].append({
                'strike': current_strike, 'volume': e_val
            })

        # Put buy: G=name(6), H=volume(7)
        g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''
        h_val = safe_num(row[7].value if len(row) > 7 else None)
        if g_val and h_val > 0:
            participants[g_val]['put_buy'] += h_val
            participants[g_val]['put_buy_detail'].append({
                'strike': current_strike, 'volume': h_val
            })

        # Call sell: N=name(13), O=volume(14)
        n_val = str(row[13].value).strip() if len(row) > 13 and row[13].value else ''
        o_val = safe_num(row[14].value if len(row) > 14 else None)
        if n_val and o_val > 0:
            participants[n_val]['call_sell'] += o_val
            participants[n_val]['call_sell_detail'].append({
                'strike': current_strike, 'volume': o_val
            })

        # Call buy: Q=name(16), R=volume(17)
        q_val = str(row[16].value).strip() if len(row) > 16 and row[16].value else ''
        r_val = safe_num(row[17].value if len(row) > 17 else None)
        if q_val and r_val > 0:
            participants[q_val]['call_buy'] += r_val
            participants[q_val]['call_buy_detail'].append({
                'strike': current_strike, 'volume': r_val
            })

    # Build result
    result = {}
    for name, data in participants.items():
        cat = classify_participant(name)
        put_net = data['put_buy'] - data['put_sell']
        call_net = data['call_buy'] - data['call_sell']
        total_volume = data['put_sell'] + data['put_buy'] + data['call_sell'] + data['call_buy']

        result[name] = {
            'category': cat,
            'category_label': PARTICIPANT_LABELS.get(cat, 'その他'),
            'put_sell': data['put_sell'],
            'put_buy': data['put_buy'],
            'put_net': put_net,
            'call_sell': data['call_sell'],
            'call_buy': data['call_buy'],
            'call_net': call_net,
            'total_volume': total_volume,
            'put_sell_detail': sorted(data['put_sell_detail'], key=lambda x: -x['volume'])[:5],
            'put_buy_detail': sorted(data['put_buy_detail'], key=lambda x: -x['volume'])[:5],
            'call_sell_detail': sorted(data['call_sell_detail'], key=lambda x: -x['volume'])[:5],
            'call_buy_detail': sorted(data['call_buy_detail'], key=lambda x: -x['volume'])[:5],
        }

    return result


def build_integrated_profiles(fut_data, op_data):
    """Section ⑨-C: Integrated profiles (futures + options cross-analysis)."""
    # Merge participant names across futures and options
    all_names = set()
    if fut_data:
        for sec in fut_data.values():
            for s in sec.get('sellers', []):
                all_names.add(s['name'])
            for b in sec.get('buyers', []):
                all_names.add(b['name'])
    if op_data:
        all_names.update(op_data.keys())

    profiles = []
    for name in all_names:
        cat = classify_participant(name)

        # Futures positions
        nk_large = 0
        nk_mini = 0
        topix = 0

        if fut_data:
            for sec_key, label in [('nk225_large', 'nk_large'), ('nk225_mini', 'nk_mini'), ('topix', 'topix')]:
                sec = fut_data.get(sec_key, {})
                sell = sum(s['volume'] for s in sec.get('sellers', []) if s['name'] == name)
                buy = sum(b['volume'] for b in sec.get('buyers', []) if b['name'] == name)
                net = buy - sell  # positive = long
                if label == 'nk_large':
                    nk_large = net if (buy > 0 or sell > 0) else 0
                elif label == 'nk_mini':
                    nk_mini = net if (buy > 0 or sell > 0) else 0
                else:
                    topix = net if (buy > 0 or sell > 0) else 0

        # Options positions
        put_net = 0
        call_net = 0
        op_detail = []
        if op_data and name in op_data:
            od = op_data[name]
            put_net = od['put_net']
            call_net = od['call_net']
            # Top strike details
            for d in od['put_sell_detail'][:3]:
                op_detail.append('P%d売%d' % (d['strike'], d['volume']))
            for d in od['put_buy_detail'][:3]:
                op_detail.append('P%d買%d' % (d['strike'], d['volume']))
            for d in od['call_sell_detail'][:3]:
                op_detail.append('C%d売%d' % (d['strike'], d['volume']))
            for d in od['call_buy_detail'][:3]:
                op_detail.append('C%d買%d' % (d['strike'], d['volume']))

        # Significance: sum of absolute values
        significance = abs(nk_large) + abs(nk_mini) / 10 + abs(topix) + abs(put_net) + abs(call_net)
        if significance == 0:
            continue

        # Estimate strategy
        fut_dir = 'long' if (nk_large + nk_mini / 10) > 0 else 'short' if (nk_large + nk_mini / 10) < 0 else 'flat'
        strategy = estimate_strategy(fut_dir, put_net, call_net)

        profiles.append({
            'name': name,
            'category': cat,
            'category_label': PARTICIPANT_LABELS.get(cat, 'その他'),
            'nk225_large': nk_large,
            'nk225_mini': nk_mini,
            'topix': topix,
            'put_net': put_net,
            'call_net': call_net,
            'op_detail': ' / '.join(op_detail) if op_detail else '',
            'strategy': strategy,
            'significance': significance,
        })

    # Sort by significance descending, take top 15
    profiles.sort(key=lambda p: -p['significance'])
    return profiles[:15]


def estimate_strategy(fut_dir, put_net, call_net):
    """Estimate trading strategy from futures direction + options net."""
    if fut_dir == 'long' and call_net < 0:
        return 'カバードコール'
    if fut_dir == 'long' and put_net > 0:
        return 'プロテクティブ・プット'
    if fut_dir == 'short' and call_net < 0:
        return 'ベア（全方位弱気）'
    if put_net < 0 and call_net > 0:
        return 'リバーサル（強気転換）'
    if put_net > 0 and call_net < 0:
        return 'プロテクティブ・カラー'
    if fut_dir == 'long' and put_net > 0 and call_net > 0:
        return 'ロングボラ'
    if put_net < 0 and call_net < 0:
        return 'ショートストラングル系'
    if fut_dir == 'long':
        return 'ブル（強気）'
    if fut_dir == 'short':
        return 'ベア（弱気）'
    return 'ニュートラル'


def extract_s10(stock_vol_path):
    """Section ⑩: Stock flow by investor type (weekly .xls).
    
    TSE Prime sheet structure:
    - Col 0: Category name (自己計, 法人, 個人, 海外投資家, 投資信託)
    - Col 6: Balance (差引き) for previous week (千円, string with commas)
    - Col 10: Balance (差引き) for current week (千円, string with commas)
    - Balance appears on either 売り or 買い row (check both rows per category)
    """
    try:
        import xlrd
    except ImportError:
        return {'error': 'xlrd not installed'}

    wb = xlrd.open_workbook(stock_vol_path)
    ws = None
    for name in wb.sheet_names():
        if 'Prime' in name or 'プライム' in name:
            ws = wb.sheet_by_name(name)
            break
    if not ws:
        ws = wb.sheet_by_index(0)

    # Extract period label from header area
    period = ''
    for r in range(min(5, ws.nrows)):
        val = str(ws.cell_value(r, 0)) if ws.ncols > 0 else ''
        if '週' in val or 'week' in val.lower():
            period = val.strip()
            break

    # Categories to extract: key -> search keywords in col 0
    categories = {
        'foreigners': ['海外投資家', 'Foreigners'],
        'individuals': ['個人', 'Individuals'],
        'institutions': ['法人', 'Institutions'],
        'investment_trusts': ['投資信託', 'Investment'],
        'proprietary': ['自己計', 'Proprietary'],
    }

    result = {'period': period}
    found_rows = {}

    # Step 1: Find the starting row for each category
    for r in range(ws.nrows):
        c0 = str(ws.cell_value(r, 0)) if ws.ncols > 0 else ''
        c0_clean = c0.replace('\u3000', '').replace(' ', '')  # strip full/half-width spaces
        for key, keywords in categories.items():
            if key not in found_rows:
                for kw in keywords:
                    if kw in c0 or kw in c0_clean:
                        found_rows[key] = r
                        break

    # Step 2: For each category, check its row and next 1-2 rows for balance values
    for key, start_row in found_rows.items():
        current_balance = None
        prev_balance = None
        for r_offset in range(3):  # check up to 3 rows
            r = start_row + r_offset
            if r >= ws.nrows:
                break
            # Column 10 = current week balance
            val10 = ws.cell_value(r, 10) if ws.ncols > 10 else ''
            if val10 and str(val10).strip() and str(val10).strip() not in ('千円,%  1,000 yen, %', '差引き Balance'):
                parsed = safe_num(val10)
                if parsed != 0 or '-0' in str(val10):
                    current_balance = parsed
            # Column 6 = previous week balance
            val6 = ws.cell_value(r, 6) if ws.ncols > 6 else ''
            if val6 and str(val6).strip():
                parsed6 = safe_num(val6)
                if parsed6 != 0 or '-0' in str(val6):
                    prev_balance = parsed6

            if current_balance is not None:
                break

        if current_balance is not None:
            result[key] = {
                'raw_1000yen': current_balance,
                'oku_yen': round(current_balance / 100000, 1),
            }
            if prev_balance is not None:
                result[key]['prev_1000yen'] = prev_balance
                result[key]['prev_oku_yen'] = round(prev_balance / 100000, 1)

    return result


def extract_s11_mini_oi(wb_oi, atm):
    """Section ⑪-A: Mini options OI."""
    if '別紙2' not in wb_oi.sheetnames:
        return {}

    ws = wb_oi['別紙2']
    pat_put = re.compile(r'NK225\s*MINI\s*P(\d{6})-(\d+)')
    pat_call = re.compile(r'NK225\s*MINI\s*C(\d{6})-(\d+)')

    # Collect all mini options data
    mini_data = defaultdict(lambda: {'puts': {}, 'calls': {}})

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ''
        g_val = str(row[6].value).strip() if len(row) > 6 and row[6].value else ''

        m = pat_put.match(a_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            oi = safe_num(row[2].value if len(row) > 2 else None)
            chg = safe_num(row[3].value if len(row) > 3 else None)
            mini_data[expiry]['puts'][strike] = {'oi': oi, 'change': chg}

        m = pat_call.match(g_val)
        if m:
            expiry = m.group(1)
            strike = int(m.group(2))
            oi = safe_num(row[8].value if len(row) > 8 else None)
            chg = safe_num(row[9].value if len(row) > 9 else None)
            mini_data[expiry]['calls'][strike] = {'oi': oi, 'change': chg}

    return dict(mini_data)


def compute_otm_table(atm, vi, days_to_sq):
    """Section ⑪-B: OTM probability table."""
    if not all([atm, vi, days_to_sq]):
        return []

    atm_round = round500(atm)
    sigma = vi / 100.0
    T = days_to_sq / 250.0

    strikes = [
        {'strike': atm_round - 5000, 'type': 'put', 'label': 'Deep OTM P'},
        {'strike': atm_round - 3000, 'type': 'put', 'label': 'OTM P'},
        {'strike': atm_round - 2000, 'type': 'put', 'label': 'OTM P'},
        {'strike': atm_round - 1000, 'type': 'put', 'label': 'Near OTM P'},
        {'strike': atm_round + 2000, 'type': 'call', 'label': 'OTM C'},
        {'strike': atm_round + 3000, 'type': 'call', 'label': 'OTM C'},
        {'strike': atm_round + 5000, 'type': 'call', 'label': 'Deep OTM C'},
    ]

    results = []
    for s in strikes:
        # 3 scenarios: VI-10, current, VI+10
        probs = {}
        for vi_scenario, vi_label in [(vi - 10, 'vi_minus10'), (vi, 'vi_current'), (vi + 10, 'vi_plus10')]:
            if vi_scenario <= 0:
                vi_scenario = 1
            sig = vi_scenario / 100.0
            T_val = T
            prob = otm_probability(atm, s['strike'], sig, T_val, s['type'])
            probs[vi_label] = round(prob * 100, 1)

        # Also compute BS price at current VI
        price = bs_price(s['type'], s['strike'], atm, sigma, T)

        results.append({
            'strike': s['strike'],
            'type': s['type'],
            'label': s['label'],
            'otm_prob': probs,
            'bs_price': round(price, 1),
        })

    return results


def compute_edge_scores(s06_data, s05_data, otm_table, s09_data=None):
    """Section ⑪-C: Zone edge scores."""
    if not s06_data or 'distribution' not in s06_data:
        return []

    atm_round = s06_data.get('atm_round', 0)
    dist = {d['strike']: d for d in s06_data['distribution']}

    zones = [
        {'name': 'Deep OTM コール', 'range': (atm_round + 5000, atm_round + 10000), 'type': 'sell-c'},
        {'name': 'OTM コール', 'range': (atm_round + 2000, atm_round + 5000), 'type': 'sell-c'},
        {'name': 'ATM帯', 'range': (atm_round - 2000, atm_round + 2000), 'type': 'atm'},
        {'name': 'OTM プット', 'range': (atm_round - 5000, atm_round - 1000), 'type': 'sell-p'},
        {'name': 'Deep OTM プット', 'range': (atm_round - 10000, atm_round - 5000), 'type': 'sell-p'},
    ]

    # Max OI for normalization
    all_oi = [d['put_oi'] for d in s06_data['distribution']] + [d['call_oi'] for d in s06_data['distribution']]
    max_oi = max(all_oi) if all_oi else 1

    result = []
    for zone in zones:
        lo, hi = zone['range']

        # 1. Wall score: max OI in zone
        wall_max = 0
        wall_strike = 0
        for strike in range(lo, hi + 1, 500):
            d = dist.get(strike, {})
            oi_val = d.get('call_oi', 0) if zone['type'] == 'sell-c' else d.get('put_oi', 0)
            if oi_val > wall_max:
                wall_max = oi_val
                wall_strike = strike
        wall_score = wall_max / max_oi if max_oi else 0

        # 2. OTM probability score
        otm_score = 0
        for ot in otm_table:
            if lo <= ot['strike'] <= hi:
                prob = ot['otm_prob'].get('vi_current', 50)
                otm_score = max(otm_score, (prob - 50) / 50)  # 90% -> 0.8, 50% -> 0
        otm_score = max(0, min(1, otm_score))

        # 3. Flow change score
        flow_score = 0
        for chg in (s05_data or []):
            if lo <= chg['strike'] <= hi:
                if zone['type'] == 'sell-p' and chg['type'] == 'P' and chg['change'] > 0:
                    flow_score = 0.5
                elif zone['type'] == 'sell-c' and chg['type'] == 'C' and chg['change'] > 0:
                    flow_score = 0.5

        # 4. Participant alignment (simplified - needs s09 data)
        participant_score = 0

        total = (wall_score + otm_score + flow_score + participant_score) / 4
        stars = max(1, min(5, int(total / 0.2) + 1))

        result.append({
            'zone': zone['name'],
            'type': zone['type'],
            'range_low': lo,
            'range_high': hi,
            'wall_max_oi': wall_max,
            'wall_strike': wall_strike,
            'wall_score': round(wall_score, 3),
            'otm_score': round(otm_score, 3),
            'flow_score': round(flow_score, 3),
            'participant_score': round(participant_score, 3),
            'total_score': round(total, 3),
            'stars': stars,
        })

    return result


def build_strategy_presets(atm, vi, T):
    """Section ⑪-F / ⑫-D: Strategy presets for P&L simulator."""
    if not atm:
        return []

    atm_r = round500(atm)
    sigma = vi / 100.0 if vi else 0.25

    presets = []

    # 1. Iron Condor
    p_sell = atm_r - 2000
    p_buy = atm_r - 4000
    c_sell = atm_r + 2000
    c_buy = atm_r + 4000
    presets.append({
        'name': 'Iron Condor',
        'legs': [
            {'type': 'put', 'side': 'short', 'strike': p_sell, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('put', p_sell, atm, sigma, T), 1)},
            {'type': 'put', 'side': 'long', 'strike': p_buy, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('put', p_buy, atm, sigma, T), 1)},
            {'type': 'call', 'side': 'short', 'strike': c_sell, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('call', c_sell, atm, sigma, T), 1)},
            {'type': 'call', 'side': 'long', 'strike': c_buy, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('call', c_buy, atm, sigma, T), 1)},
        ],
    })

    # 2. Put Credit Spread
    presets.append({
        'name': 'Put Credit Spread',
        'legs': [
            {'type': 'put', 'side': 'short', 'strike': atm_r - 2000, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('put', atm_r - 2000, atm, sigma, T), 1)},
            {'type': 'put', 'side': 'long', 'strike': atm_r - 4000, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('put', atm_r - 4000, atm, sigma, T), 1)},
        ],
    })

    # 3. Call Credit Spread
    presets.append({
        'name': 'Call Credit Spread',
        'legs': [
            {'type': 'call', 'side': 'short', 'strike': atm_r + 2000, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('call', atm_r + 2000, atm, sigma, T), 1)},
            {'type': 'call', 'side': 'long', 'strike': atm_r + 4000, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('call', atm_r + 4000, atm, sigma, T), 1)},
        ],
    })

    # 4. Short Strangle
    presets.append({
        'name': 'Short Strangle + 先物ヘッジ',
        'legs': [
            {'type': 'put', 'side': 'short', 'strike': atm_r - 2000, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('put', atm_r - 2000, atm, sigma, T), 1)},
            {'type': 'call', 'side': 'short', 'strike': atm_r + 2000, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('call', atm_r + 2000, atm, sigma, T), 1)},
            {'type': 'futures', 'side': 'short', 'entry': atm, 'qty': 0, 'multiplier': 100},
        ],
    })

    # 5. P売り + miniヘッジ
    presets.append({
        'name': 'P売り + mini先物ヘッジ',
        'legs': [
            {'type': 'put', 'side': 'short', 'strike': atm_r - 2000, 'qty': 1, 'multiplier': 1000,
             'premium': round(bs_price('put', atm_r - 2000, atm, sigma, T), 1)},
            {'type': 'futures', 'side': 'short', 'entry': atm_r - 1500, 'qty': 3, 'multiplier': 100},
        ],
    })

    # 6. Mini weekly theta harvest
    presets.append({
        'name': 'ミニ週次θハーベスト',
        'legs': [
            {'type': 'put', 'side': 'short', 'strike': atm_r - 1500, 'qty': 10, 'multiplier': 100,
             'premium': round(bs_price('put', atm_r - 1500, atm, sigma, max(T * 0.3, 5 / 250)), 1)},
            {'type': 'call', 'side': 'short', 'strike': atm_r + 1500, 'qty': 10, 'multiplier': 100,
             'premium': round(bs_price('call', atm_r + 1500, atm, sigma, max(T * 0.3, 5 / 250)), 1)},
        ],
    })

    return presets


# ============================================================
# Derived Indicators
# ============================================================

def compute_indicators(data):
    """Compute derived trading indicators from extracted data."""
    indicators = {}
    s03 = data.get('s03', {})
    s04 = data.get('s04', {})
    s06 = data.get('s06', {})
    s01 = data.get('s01', {})
    meta = data.get('metadata', {})

    # --- Put/Call Ratio ---
    lg = s03.get('large', {})
    pv = lg.get('put_volume', 0)
    cv = lg.get('call_volume', 0)
    indicators['pcr_volume'] = round(pv / cv, 2) if cv else None

    lg_oi = s04.get('large', {})
    p_oi = lg_oi.get('put_total_oi', 0)
    c_oi = lg_oi.get('call_total_oi', 0)
    indicators['pcr_oi'] = round(p_oi / c_oi, 2) if c_oi else None

    # PCR interpretation
    pcr = indicators.get('pcr_volume')
    if pcr is not None:
        if pcr > 1.3:
            indicators['pcr_signal'] = '強い弱気（逆張り買いシグナル）'
        elif pcr > 1.0:
            indicators['pcr_signal'] = 'やや弱気'
        elif pcr > 0.7:
            indicators['pcr_signal'] = 'ニュートラル'
        elif pcr > 0.5:
            indicators['pcr_signal'] = 'やや強気'
        else:
            indicators['pcr_signal'] = '強い強気（逆張り売りシグナル）'

    # --- Max Pain ---
    dist = s06.get('distribution', [])
    if dist:
        min_pain = None
        max_pain_strike = None
        for target in dist:
            K = target['strike']
            total_pain = 0
            for d in dist:
                # Pain for put holders if SQ = K
                if d['put_oi'] > 0:
                    put_itm = max(d['strike'] - K, 0)
                    total_pain += put_itm * d['put_oi']
                # Pain for call holders if SQ = K
                if d['call_oi'] > 0:
                    call_itm = max(K - d['strike'], 0)
                    total_pain += call_itm * d['call_oi']
            if min_pain is None or total_pain < min_pain:
                min_pain = total_pain
                max_pain_strike = K

        indicators['max_pain'] = max_pain_strike
        indicators['max_pain_total'] = min_pain

        # Distance from ATM
        atm = meta.get('atm')
        if atm and max_pain_strike:
            indicators['max_pain_diff'] = max_pain_strike - atm

    # --- Futures Basis ---
    nikkei = s01.get('nikkei_close')
    atm = meta.get('atm')
    if nikkei and atm:
        basis = atm - nikkei
        indicators['basis'] = round(basis)
        indicators['basis_pct'] = round(basis / nikkei * 100, 2)
        if basis > 0:
            indicators['basis_signal'] = 'コンタンゴ（先物プレミアム）'
        else:
            indicators['basis_signal'] = 'バックワーデーション（先物ディスカウント）'

    # --- Wall Changes Summary ---
    if dist:
        reinforced = []
        weakened = []
        for d in dist:
            # Put walls
            if d['put_oi'] > 3000 and d['put_change'] > 200:
                reinforced.append({'strike': d['strike'], 'type': 'P', 'oi': d['put_oi'], 'change': d['put_change']})
            elif d['put_oi'] > 3000 and d['put_change'] < -200:
                weakened.append({'strike': d['strike'], 'type': 'P', 'oi': d['put_oi'], 'change': d['put_change']})
            # Call walls
            if d['call_oi'] > 3000 and d['call_change'] > 200:
                reinforced.append({'strike': d['strike'], 'type': 'C', 'oi': d['call_oi'], 'change': d['call_change']})
            elif d['call_oi'] > 3000 and d['call_change'] < -200:
                weakened.append({'strike': d['strike'], 'type': 'C', 'oi': d['call_oi'], 'change': d['call_change']})

        reinforced.sort(key=lambda x: -x['change'])
        weakened.sort(key=lambda x: x['change'])
        indicators['walls_reinforced'] = reinforced[:5]
        indicators['walls_weakened'] = weakened[:5]

    return indicators


# ============================================================
# Main pipeline
# ============================================================

def run(args):
    data_dir = args.dir
    files = detect_files(data_dir)

    print('[extract.py] Detected files:')
    for k, v in files.items():
        if k != 'date_str':
            print('  %s: %s' % (k, os.path.basename(v)))

    # Parse date
    date_str = files.get('date_str', datetime.now().strftime('%Y%m%d'))
    analysis_date = datetime.strptime(date_str, '%Y%m%d')
    weekday_ja = WEEKDAY_JA[analysis_date.weekday()]

    # Determine SQ
    major_y, major_m = next_major_month(analysis_date)
    sq = sq_date(major_y, major_m)
    days_to_sq = business_days_between(analysis_date, sq)

    output = {
        'metadata': {
            'date': date_str,
            'date_formatted': '%d/%d/%d（%s）' % (analysis_date.year, analysis_date.month, analysis_date.day, weekday_ja),
            'sq_date': sq.strftime('%Y-%m-%d'),
            'sq_label': '%d月限SQ（%s）' % (major_m, sq.strftime('%m/%d')),
            'days_to_sq': days_to_sq,
            'major_month': '%04d%02d' % (major_y, major_m),
            'files_found': {k: os.path.basename(v) for k, v in files.items() if k != 'date_str'},
        }
    }

    # Load workbooks
    wb_oi = None
    wb_market = None
    wb_jnet = None
    wb_fut = None
    wb_op = None

    if 'open_interest' in files:
        wb_oi = openpyxl.load_workbook(files['open_interest'], data_only=True)
        print('[extract.py] Loaded open_interest: sheets =', wb_oi.sheetnames)

    if 'market_data' in files:
        wb_market = openpyxl.load_workbook(files['market_data'], data_only=True)
        print('[extract.py] Loaded market_data: sheets =', wb_market.sheetnames)

    if 'jnet' in files:
        wb_jnet = openpyxl.load_workbook(files['jnet'], data_only=True)
        print('[extract.py] Loaded J-NET: sheets =', wb_jnet.sheetnames)

    if 'fut_participants' in files:
        wb_fut = openpyxl.load_workbook(files['fut_participants'], data_only=True)
        print('[extract.py] Loaded futures participants')

    if 'op_participants' in files:
        wb_op = openpyxl.load_workbook(files['op_participants'], data_only=True)
        print('[extract.py] Loaded options participants')

    # Extract ATM + OHLC
    atm = None
    ohlc = {}
    if wb_market:
        atm = extract_atm(wb_market)
        print('[extract.py] ATM =', atm)
        ohlc = extract_ohlc_pivot(wb_market)
        if ohlc:
            print('[extract.py] OHLC: O=%s H=%s L=%s C=%s PP=%s' % (
                ohlc.get('open'), ohlc.get('high'), ohlc.get('low'), ohlc.get('close'), ohlc.get('pivot')))
    output['metadata']['atm'] = atm
    output['metadata']['ohlc'] = ohlc if wb_market else {}

    # Nikkei / VI (from args, or fallback to Excel)
    nikkei = args.nikkei
    vi = args.vi
    if vi is None and wb_market:
        vi = extract_vi_from_excel(wb_market)
        if vi:
            print('[extract.py] VI from Excel (先物清算値): %.1f' % vi)
    if nikkei is None and atm:
        nikkei = float(atm)
        print('[extract.py] Nikkei fallback to ATM: %s' % atm)
    output['metadata']['nikkei_close'] = nikkei
    output['metadata']['vi'] = vi

    # ① Nikkei / VI
    output['s01'] = extract_s01(nikkei, vi, atm)
    print('[extract.py] ① done')

    # ② Futures OI
    if wb_oi:
        output['s02'] = extract_s02(wb_oi)
        print('[extract.py] ② done')
    else:
        output['s02'] = {'error': 'open_interest not found'}

    # ③ Options volume
    if wb_market:
        output['s03'] = extract_s03(wb_market)
        print('[extract.py] ③ done')
    else:
        output['s03'] = {'error': 'market_data not found'}

    # ④ Options OI changes
    if wb_oi:
        output['s04'] = extract_s04(wb_oi)
        print('[extract.py] ④ done')
    else:
        output['s04'] = {'error': 'open_interest not found'}

    # ⑤ Important OI changes
    if wb_oi:
        output['s05'] = extract_s05(wb_oi)
        print('[extract.py] ⑤ done: %d entries' % len(output['s05']))
    else:
        output['s05'] = []

    # ⑥ OI distribution
    if wb_oi and atm:
        output['s06'] = extract_s06(wb_oi, atm)
        print('[extract.py] ⑥ done')
    else:
        output['s06'] = {'error': 'ATM or open_interest not available'}

    # ⑦ J-NET trades
    if wb_jnet:
        output['s07'] = extract_s07(wb_jnet)
        print('[extract.py] ⑦ done: %d trades' % len(output['s07']))
    else:
        output['s07'] = []

    # ⑨ Participant analysis (with caching)
    cache_s09_path = os.path.join(data_dir, 'cache_s09.json')
    fut_data = None
    op_data = None

    if wb_fut:
        fut_data = extract_s09_futures(wb_fut)
        print('[extract.py] ⑨-A done')

    if wb_op:
        op_data = extract_s09_options(wb_op)
        print('[extract.py] ⑨-B done: %d participants' % len(op_data))

    if fut_data or op_data:
        profiles = build_integrated_profiles(fut_data, op_data)
        output['s09'] = {
            'futures': fut_data or {},
            'options': op_data or {},
            'profiles': profiles,
            'source': 'live',
            'data_date': date_str,
        }
        print('[extract.py] ⑨-C done: %d profiles' % len(profiles))

        # Save to cache
        try:
            with open(cache_s09_path, 'w', encoding='utf-8') as f:
                json.dump(output['s09'], f, ensure_ascii=False, indent=2)
            print('[extract.py] ⑨ cached to %s' % cache_s09_path)
        except Exception as e:
            print('[extract.py] ⑨ cache write failed: %s' % e)

    elif os.path.exists(cache_s09_path):
        # Load from cache
        try:
            with open(cache_s09_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            cached['source'] = 'cache'
            cached_date = cached.get('data_date', '?')
            output['s09'] = cached
            print('[extract.py] ⑨ loaded from cache (%s時点・参考)' % cached_date)
        except Exception as e:
            print('[extract.py] ⑨ cache read failed: %s' % e)
            output['s09'] = {'error': 'No weekly data and cache unreadable'}
    else:
        output['s09'] = {'error': 'No weekly participant files'}

    # ⑩ Stock flow (with caching)
    cache_s10_path = os.path.join(data_dir, 'cache_s10.json')

    if 'stock_vol' in files:
        output['s10'] = extract_s10(files['stock_vol'])
        output['s10']['source'] = 'live'
        output['s10']['data_date'] = date_str
        print('[extract.py] ⑩ done')

        # Save to cache
        try:
            with open(cache_s10_path, 'w', encoding='utf-8') as f:
                json.dump(output['s10'], f, ensure_ascii=False, indent=2)
            print('[extract.py] ⑩ cached')
        except:
            pass

    elif os.path.exists(cache_s10_path):
        try:
            with open(cache_s10_path, 'r', encoding='utf-8') as f:
                cached = json.load(f)
            cached['source'] = 'cache'
            output['s10'] = cached
            print('[extract.py] ⑩ loaded from cache (%s時点)' % cached.get('data_date', '?'))
        except:
            output['s10'] = {'skipped': True}
    else:
        output['s10'] = {'skipped': True}

    # ⑪ Strategy map
    T = days_to_sq / 250.0 if days_to_sq else 0
    s11 = {}

    if wb_oi and atm:
        s11['mini_oi'] = extract_s11_mini_oi(wb_oi, atm)
        print('[extract.py] ⑪-A done: %d expiries' % len(s11['mini_oi']))

    if atm and vi and days_to_sq:
        s11['otm_table'] = compute_otm_table(atm, vi, days_to_sq)
        s11['edge_scores'] = compute_edge_scores(
            output.get('s06'), output.get('s05'), s11.get('otm_table', []),
            output.get('s09')
        )
        s11['presets'] = build_strategy_presets(atm, vi, T)
        print('[extract.py] ⑪-B/C/F done')

    output['s11'] = s11

    # Derived indicators (PCR, Max Pain, Basis, Wall changes)
    output['indicators'] = compute_indicators(output)
    ind = output['indicators']
    print('[extract.py] Indicators: PCR=%.2f MaxPain=%s Basis=%s' % (
        ind.get('pcr_volume') or 0, ind.get('max_pain', '-'), ind.get('basis', '-')))

    # Write output
    out_path = args.out
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print('\n[extract.py] Output written to: %s' % out_path)
    print('[extract.py] JSON size: %.1f KB' % (os.path.getsize(out_path) / 1024))

    return output


# ============================================================
# CLI
# ============================================================

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='JPX Market Analysis - Data Extractor')
    parser.add_argument('--dir', default='.', help='Directory containing JPX Excel files')
    parser.add_argument('--out', default='data.json', help='Output JSON path')
    parser.add_argument('--nikkei', type=float, default=None, help='Nikkei 225 cash close')
    parser.add_argument('--vi', type=float, default=None, help='Nikkei VI close')
    args = parser.parse_args()

    run(args)
